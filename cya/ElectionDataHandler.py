import pandas as pd
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from matplotlib.widgets import Slider

def fill_blank_cells_in_column(input_excel_file):
    # Load the workbook and get the active sheet
    wb = openpyxl.load_workbook(input_excel_file)
    sheet = wb.active  # Get the currently active sheet

    # Start from the second row (skip the header)
    current_value = None  # To keep track of the current value to fill blanks
    
    # Iterate over all rows in column A (starting from row 2, since row 1 is headers)
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=1)  # Column A is column 1
        
        if cell.value is not None:  # If the cell is not empty, update current_value
            current_value = cell.value
        elif current_value is not None:  # If the cell is empty and we have a current_value
            cell.value = current_value  # Fill the blank with the current value

    # Save the changes to the Excel file
    wb.save(input_excel_file)

    print(f"Processed active sheet and filled blanks in column A of {input_excel_file}")

def fill_column_b(input_excel_file):
    # Load the workbook and get the active sheet
    wb = openpyxl.load_workbook(input_excel_file)
    sheet = wb.active  # Get the currently active sheet

    # Iterate over all rows in column B, starting from the second row (since row 1 is header)
    row = 2  # Start from row 2 (ignoring header)
    
    while row <= sheet.max_row:
        current_cell = sheet.cell(row=row, column=2)  # Column B is column 2
        next_cell = sheet.cell(row=row + 1, column=2)  # The cell in the next row (below the current one)

        # Stop processing if we encounter 'E'
        if current_cell.value == 'E':
            print(f"Encountered 'E' at row {row}. Stopping process.")
            break

        # If the current cell is 'R' and the next cell is blank, set it to 'D'
        if current_cell.value == 'R' and next_cell.value is None:
            next_cell.value = 'D'  # Set the next row to 'D'
            row += 2  # Skip the next row (which was just filled with 'D')

        # If the current cell is 'D' and the next cell is blank, set it to 'R'
        elif current_cell.value == 'D' and next_cell.value is None:
            next_cell.value = 'R'  # Set the next row to 'R'
            row += 2  # Skip the next row (which was just filled with 'R')

        else:
            row += 1  # If no 'R' or 'D', just move to the next row

    # After applying the R-D condition, re-check the column for non-blank values
    # and set remaining blanks to 'I' while re-applying the R-D logic.
    for row in range(2, sheet.max_row + 1):
        current_cell = sheet.cell(row=row, column=2)  # Column B is column 2
        if current_cell.value is None:  # If the cell is blank
            current_cell.value = 'I'  # Set it to 'I'

    # Save the changes to the Excel file
    wb.save(input_excel_file)

    print(f"Processed active sheet and filled column B in {input_excel_file}")





def process_votes(input_excel_file):
    # Load the Excel file using openpyxl to modify sheets
    wb = load_workbook(input_excel_file)

    # If the "Output" sheet exists, delete it
    if 'Output' in wb.sheetnames:
        del wb['Output']

    # Initialize a list to collect the data for all sheets
    all_output_data = []

    # Load the data and process it for each sheet in the workbook
    excel_file = pd.ExcelFile(input_excel_file)
    
    # Iterate through each sheet in the workbook
    for sheet_name in excel_file.sheet_names:
        print(f"Processing sheet: {sheet_name}")  # Debugging line to see which sheet is being processed
        
        # Read the data from the sheet, using only the first three columns (A, B, C)
        df = pd.read_excel(excel_file, sheet_name=sheet_name, usecols=[0, 1, 2], header=None)
        
        # Check if the sheet is empty
        if df.empty:
            print(f"Warning: Sheet '{sheet_name}' is empty or does not contain valid data.")
            continue  # Skip this sheet if it's empty

        # Rename the columns for easier access
        df.columns = ['State Abbreviation', 'Party', 'Votes']
        
        # Ensure the Votes column is numeric (in case there are any non-numeric values)
        df['Votes'] = pd.to_numeric(df['Votes'], errors='coerce').fillna(0).astype(int)
        
        # Initialize empty dictionaries for each state
        state_votes = {}
        
        # Process each row in the data
        for _, row in df.iterrows():
            state = row['State Abbreviation']
            party = row['Party']
            votes = row['Votes']
            
            # Skip rows with invalid state abbreviations (either blank or not 2 letters long)
            if not isinstance(state, str) or len(state.strip()) != 2:
                print(f"Skipping row with invalid state abbreviation: '{state}'")
                continue  # Skip this row
            
            # Skip rows where the Party column is blank or NaN
            if pd.isna(party) or not party.strip():
                print(f"Skipping row with blank or invalid party value for state: '{state}'")
                continue  # Skip this row
            
            # Initialize state record if not already present
            if state not in state_votes:
                state_votes[state] = {'Republican': 0, 'Democrat': 0, 'Other': 0}
            
            # Add the votes to the correct bucket based on the party
            if party == 'R':
                state_votes[state]['Republican'] += votes
            elif party == 'D':
                state_votes[state]['Democrat'] += votes
            else:
                state_votes[state]['Other'] += votes
        
        # Create the output data for this sheet
        for state, vote_data in state_votes.items():
            all_output_data.append({
                'Year': sheet_name,  # Use the sheet name as the year
                'State': state,
                'Republican': vote_data['Republican'],
                'Democrat': vote_data['Democrat'],
                'Other': vote_data['Other']
            })
    
    # Check if any data has been collected
    if not all_output_data:
        print("Warning: No data was collected for the output sheet.")
        return
    
    # Convert the collected data into a DataFrame
    output_df = pd.DataFrame(all_output_data)
    
    # Write the output data to the "Output" sheet
    with pd.ExcelWriter(input_excel_file, engine='openpyxl', mode='a') as writer:
        # Adding the new "Output" sheet to the workbook
        output_df.to_excel(writer, sheet_name='Output', index=False)

    # No need to call wb.save(input_excel_file) anymore as ExcelWriter saves the file
    print(f"Processed votes and added Output tab to {input_excel_file}")





def generate_party_chart(input_excel_file, state_abbreviation, start_year, end_year, parties):
    # Load the Excel file
    df = pd.read_excel(input_excel_file, sheet_name='Output')
    
    # Filter data based on input conditions
    filtered_df = df[
        (df['State'] == state_abbreviation) & 
        (df['Year'].astype(int) >= start_year) & 
        (df['Year'].astype(int) <= end_year)
    ]
    
    # If there's no data for the given state, year range
    if filtered_df.empty:
        print(f"No data found for {state_abbreviation} between {start_year} and {end_year}.")
        return

    # Create a DataFrame for the selected parties
    # Only include the columns for the parties selected
    selected_columns = ['Year'] + parties
    filtered_df = filtered_df[selected_columns]

    # If the requested parties are not in the columns, print an error
    for party in parties:
        if party not in filtered_df.columns:
            print(f"Warning: No data found for the party '{party}' in the selected sheet.")
            return

    # Plotting the bar chart for selected parties
    ax = filtered_df.set_index('Year')[parties].plot(kind='bar', figsize=(10, 6), width=0.8, colormap='Set2')

    # Customize the chart
    ax.set_title(f"Votes in {state_abbreviation} ({start_year}-{end_year})", fontsize=16)
    ax.set_xlabel('Year', fontsize=12)
    ax.set_ylabel('Votes', fontsize=12)
    ax.set_xticklabels([str(x) for x in filtered_df['Year']], rotation=45, ha='right')
    ax.legend(title='Party', loc='upper left')

    plt.tight_layout()
    plt.show()



def generate_all_states_chart(input_excel_file, start_year, end_year, parties, mode='A'):
    # Load the data from the Output sheet
    df = pd.read_excel(input_excel_file, sheet_name='Output')

    # Filter data for the selected year range
    filtered_df = df[
        (df['Year'].astype(int) >= start_year) & 
        (df['Year'].astype(int) <= end_year)
    ]

    # Get the unique states from the data
    states = filtered_df['State'].unique()

    # Calculate the sorting criteria based on mode
    if mode in ['D', 'R']:
        differences = []
        for state in states:
            state_df = filtered_df[filtered_df['State'] == state]
            target_party = 'Democrat' if mode == 'D' else 'Republican'
            
            # Get the 2020 vote count for the target party
            votes_2020 = state_df[state_df['Year'] == 2020][target_party].values
            if votes_2020.size > 0:
                # Calculate the highest vote count from non-2020 years
                non_2020_max = state_df[state_df['Year'] != 2020][target_party].max()
                if pd.notna(non_2020_max):  # Check if there's a valid non-2020 max
                    difference = votes_2020[0] - non_2020_max
                    differences.append((state, difference))
                else:
                    differences.append((state, 0))  # If no non-2020 data, set diff to 0
            else:
                differences.append((state, 0))  # If no 2020 data, set diff to 0
        
        # Sort states based on the computed differences
        sorted_states = [state for state, _ in sorted(differences, key=lambda x: x[1], reverse=True)]
    else:
        sorted_states = states  # No sorting for "A" mode

    # Set up the figure and axis for the plot
    fig, ax = plt.subplots(figsize=(10, 6))
    plt.subplots_adjust(bottom=0.25)  # To make space for the slider
    
    # Set up the slider axis
    ax_slider = plt.axes([0.1, 0.01, 0.8, 0.03], facecolor='lightgoldenrodyellow')
    slider = Slider(ax_slider, 'State', 0, len(sorted_states) - 1, valinit=0, valstep=1)

    # Define custom colors for the parties
    party_colors = {
        'Republican': 'red',
        'Democrat': 'blue',
        'Other': 'yellow'
    }

    # Function to update the chart based on the selected state
    def update(val):
        state_idx = int(slider.val)
        state = sorted_states[state_idx]
        
        # Filter data for the selected state
        state_df = filtered_df[filtered_df['State'] == state]
        
        # Plotting the bar chart for selected parties with custom colors
        ax.clear()  # Clear the previous plot
        state_df.set_index('Year')[parties].plot(kind='bar', ax=ax, width=0.8, color=[party_colors.get(party, 'gray') for party in parties])
        
        # Determine the target party for overvote (Democrat or Republican) if in "D" or "R" mode
        target_party = None
        difference = None
        if mode == 'D':
            target_party = 'Democrat'
        elif mode == 'R':
            target_party = 'Republican'
        
        # If a target party is set, calculate the difference and add the lines
        if target_party:
            # Get the 2020 vote count for the target party
            votes_2020 = state_df[state_df['Year'] == 2020][target_party].values
            if votes_2020.size > 0:
                # Calculate the highest vote count from non-2020 years
                non_2020_max = state_df[state_df['Year'] != 2020][target_party].max()
                if pd.notna(non_2020_max):  # Check if there's a valid non-2020 max
                    difference = votes_2020[0] - non_2020_max
                    ax.axhline(y=votes_2020[0], color='purple', linestyle='--', linewidth=2, label=f'{target_party} 2020')
                    ax.axhline(y=non_2020_max, color='gray', linestyle='--', linewidth=2, label=f'{target_party} (Highest non-2020)')
        
        # Set the title, labels, and legend
        ax.set_title(f"Votes in {state} ({start_year}-{end_year})", fontsize=16)
        ax.set_xlabel('Year', fontsize=12)
        ax.set_ylabel('Votes', fontsize=12)
        ax.set_xticklabels([str(x) for x in state_df['Year']], rotation=45, ha='right')
        
        # Add the calculated difference to the legend with commas if available
        if difference is not None:
            formatted_difference = f"{difference:,.0f}"  # Format with commas
            ax.legend(title=f'{target_party} 2020 Overvote: {formatted_difference}', loc='lower center', bbox_to_anchor=(0.5, 1.05), ncol=3)
        else:
            ax.legend(title='Party', loc='lower center', bbox_to_anchor=(0.5, 1.05), ncol=3)
        
        plt.draw()


    # Initialize the plot with the first state
    update(0)
    
    # Attach the update function to the slider
    slider.on_changed(update)

    # Show the plot
    plt.show()

# Usage
input_excel_file = "C:/Users/Judson/Desktop/Scirpts/Election/!Combined.xlsx"



# Example of how to call the function
#generate_party_chart(input_excel_file, state_abbreviation='AL', start_year=2008, end_year=2024, parties=['Republican', 'Democrat'])
generate_all_states_chart(input_excel_file, start_year=2000, end_year=2024, parties=['Republican', 'Democrat', 'Other'], mode="D")





#process_votes(input_excel_file)
#fill_blank_cells_in_column(input_excel_file)
#fill_column_b(input_excel_file)